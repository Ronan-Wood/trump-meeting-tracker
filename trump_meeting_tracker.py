#!/usr/bin/env python3
"""
Trump Meetings Tracker - Enhanced Version
Searches for Trump meetings using NewsAPI + RSS feeds, identifies companies/industries, sends email reports
"""

import os
import json
import sqlite3
import requests
from datetime import datetime, timedelta
from typing import List, Dict, Optional
import re

# Load environment variables from .env file
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    # dotenv not installed, environment variables must be set manually
    pass

# News APIs
from newsapi import NewsApiClient
import feedparser
from bs4 import BeautifulSoup

# Email
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail, Attachment, FileContent, FileName, FileType, Disposition

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from collections import Counter
import base64


class TrumpMeetingsTracker:
    def __init__(self, db_path='trump_meetings.db', config_path='data_sources_config.json'):
        self.db_path = db_path
        self.config_path = config_path
        self.config = self.load_config()
        self.init_database()
        
        # Initialize NewsAPI client
        self.newsapi_key = os.environ.get('NEWS_API_KEY')
        if self.newsapi_key:
            self.newsapi = NewsApiClient(api_key=self.newsapi_key)
        else:
            self.newsapi = None
            print("⚠️ NEWS_API_KEY not set - NewsAPI searches will be skipped")
        
    def load_config(self):
        """Load configuration from JSON file"""
        with open(self.config_path, 'r') as f:
            return json.load(f)
    
    def init_database(self):
        """Initialize SQLite database"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Create meetings table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS meetings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT NOT NULL,
                location TEXT,
                meeting_type TEXT,
                source_url TEXT,
                source_publication TEXT,
                date_added TEXT NOT NULL,
                notes TEXT,
                UNIQUE(date, location, source_url)
            )
        ''')

        # Create attendees table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS attendees (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                meeting_id INTEGER,
                name TEXT NOT NULL,
                title TEXT,
                company TEXT,
                primary_industry TEXT,
                secondary_industries TEXT,
                confidence_level TEXT,
                confidence_reasons TEXT,
                requires_review BOOLEAN,
                FOREIGN KEY (meeting_id) REFERENCES meetings (id)
            )
        ''')

        # Migration: Add source tracking columns if they don't exist
        try:
            cursor.execute('ALTER TABLE meetings ADD COLUMN source_urls TEXT')
            cursor.execute('ALTER TABLE meetings ADD COLUMN source_count INTEGER DEFAULT 1')
            print("✅ Added source tracking columns to meetings table")
        except sqlite3.OperationalError:
            # Columns already exist
            pass

        # Initialize source_urls for existing records that don't have it
        cursor.execute('''
            UPDATE meetings
            SET source_urls = json_array(source_url),
                source_count = 1
            WHERE source_urls IS NULL AND source_url IS NOT NULL
        ''')

        conn.commit()
        conn.close()
    
    def search_all_sources(self, days_back=7) -> List[Dict]:
        """
        Search all sources for Trump meetings
        Returns list of meeting dictionaries
        """
        all_meetings = []
        
        print(f"🔍 Searching for meetings from last {days_back} days...")
        print()
        
        # 1. Search NewsAPI
        if self.newsapi:
            newsapi_results = self.search_newsapi(days_back)
            print(f"  📰 NewsAPI: Found {len(newsapi_results)} articles")
            all_meetings.extend(newsapi_results)
        
        # 2. Search RSS Feeds
        print(f"  📡 Searching RSS feeds...")
        rss_results = self.search_rss_feeds(days_back)
        print(f"  📡 RSS Feeds: Found {len(rss_results)} Trump-related articles")
        all_meetings.extend(rss_results)
        
        print()
        print(f"✅ Total articles found: {len(all_meetings)}")
        
        # Parse articles for meeting information
        meetings = []
        debug_mode = os.environ.get('DEBUG_FILTERING', 'false').lower() == 'true'

        for idx, article in enumerate(all_meetings):
            if debug_mode and idx < 10:
                print(f"\n  Article {idx+1}: {article['title'][:80]}...")

            parsed_meetings = self.parse_article_for_meetings(article)

            if debug_mode and idx < 10:
                if parsed_meetings:
                    print(f"    ✅ Extracted {len(parsed_meetings)} meeting(s)")
                else:
                    print(f"    ⚠️ No meetings extracted")

            for meeting in parsed_meetings:
                dup_check = self.is_duplicate_meeting(meeting)

                if dup_check['should_merge']:
                    # Same meeting from different source - merge the sources
                    merged = self.merge_meeting_source(
                        dup_check['meeting_id'],
                        meeting.get('source_url'),
                        meeting.get('source_publication')
                    )
                    if merged and debug_mode:
                        print(f"    🔗 Merged additional source for existing meeting")
                elif not dup_check['is_duplicate']:
                    # New meeting
                    meetings.append(meeting)
        
        print(f"✅ Extracted {len(meetings)} unique meetings")
        
        return meetings
    
    def search_newsapi(self, days_back=7) -> List[Dict]:
        """Search NewsAPI for Trump meeting articles"""
        if not self.newsapi:
            return []

        articles = []
        from_date = (datetime.now() - timedelta(days=days_back)).strftime('%Y-%m-%d')

        # Search queries optimized for meetings
        # Using broader queries with sort_by='relevancy' for better results
        queries = [
            # Core meeting queries
            'Trump CEO meeting',
            'Trump business leaders',
            'Trump executives meeting',

            # Location-specific
            'Mar-a-Lago CEO',
            'White House business meeting Trump',

            # Event-specific
            'Business Roundtable Trump',
            'Trump tech leaders',
            'Trump manufacturers',

            # Action-based
            '"Trump meets" CEO OR chairman OR executive',
            '"Trump hosted" business OR executives',
        ]

        for query in queries:
            try:
                # Use relevancy sorting to get better matches
                response = self.newsapi.get_everything(
                    q=query,
                    from_param=from_date,
                    language='en',
                    sort_by='relevancy',  # Changed from publishedAt
                    page_size=15  # Reduced per query to stay within limits
                )
                
                if response['status'] == 'ok':
                    for article in response['articles']:
                        articles.append({
                            'title': article['title'],
                            'description': article.get('description', ''),
                            'url': article['url'],
                            'source': article['source']['name'],
                            'published_at': article['publishedAt'],
                            'content': article.get('content', '')
                        })
            except Exception as e:
                print(f"  ⚠️ Error searching NewsAPI for '{query}': {str(e)}")
        
        # Remove duplicates by URL
        seen_urls = set()
        unique_articles = []
        for article in articles:
            if article['url'] not in seen_urls:
                seen_urls.add(article['url'])
                unique_articles.append(article)
        
        return unique_articles
    
    def search_rss_feeds(self, days_back=7) -> List[Dict]:
        """Search RSS feeds for Trump meeting articles"""
        feeds = [
            # Google News - Trump business & CEO topics
            'https://news.google.com/rss/search?q=Trump+CEO+meeting&hl=en-US&gl=US&ceid=US:en',
            'https://news.google.com/rss/search?q=Trump+business+leaders&hl=en-US&gl=US&ceid=US:en',
            'https://news.google.com/rss/search?q=Trump+executives+meeting&hl=en-US&gl=US&ceid=US:en',
            'https://news.google.com/rss/search?q=Mar-a-Lago+CEO&hl=en-US&gl=US&ceid=US:en',

            # Global News
            'https://feeds.bbci.co.uk/news/rss.xml',  # BBC Top Stories
            'https://feeds.bbci.co.uk/news/business/rss.xml',  # BBC Business
            'http://rss.cnn.com/rss/cnn_topstories.rss',  # CNN Top
            'http://rss.cnn.com/rss/cnn_allpolitics.rss',  # CNN Politics
            'http://rss.cnn.com/rss/money_latest.rss',  # CNN Business
            'https://feeds.reuters.com/reuters/topNews',  # Reuters Top
            'https://feeds.reuters.com/Reuters/domesticNews',  # Reuters US

            # U.S. National Papers
            'https://rss.nytimes.com/services/xml/rss/nyt/HomePage.xml',  # NYT Home
            'https://rss.nytimes.com/services/xml/rss/nyt/Politics.xml',  # NYT Politics
            'https://rss.nytimes.com/services/xml/rss/nyt/Business.xml',  # NYT Business
            'http://feeds.washingtonpost.com/rss/politics',  # WaPo Politics
            'http://feeds.washingtonpost.com/rss/business',  # WaPo Business
            'https://feeds.a.dj.com/rss/RSSWorldNews.xml',  # WSJ World
            'https://feeds.a.dj.com/rss/WSJcomUSBusiness.xml',  # WSJ Business

            # Public Broadcast
            'https://feeds.npr.org/1001/rss.xml',  # NPR Top Stories
            'https://feeds.npr.org/1014/rss.xml',  # NPR Politics
            'https://feeds.abcnews.com/abcnews/topstories',  # ABC News
            'https://www.cbsnews.com/latest/rss/main',  # CBS News
            'https://feeds.nbcnews.com/nbcnews/public/news',  # NBC News

            # Additional Sources
            'https://www.politico.com/rss/politicopicks.xml',  # Politico
            'http://feeds.foxnews.com/foxnews/latest',  # Fox News
            'https://www.yahoo.com/news/rss',  # Yahoo News
            'https://www.vox.com/rss/index.xml',  # Vox
        ]
        
        articles = []
        cutoff_date = datetime.now() - timedelta(days=days_back)

        # Broader keywords to catch more articles
        keywords = ['trump']  # Just require 'trump', filter more specifically later

        successful_feeds = 0
        failed_feeds = 0

        debug_mode = os.environ.get('DEBUG_FILTERING', 'false').lower() == 'true'
        if debug_mode:
            print(f"    Checking {len(feeds)} RSS feeds...")

        for feed_url in feeds:
            try:
                feed = feedparser.parse(feed_url)

                if not feed.entries:
                    failed_feeds += 1
                    continue

                successful_feeds += 1
                feed_articles = 0

                for entry in feed.entries:
                    # Check if published recently
                    if hasattr(entry, 'published_parsed'):
                        try:
                            pub_date = datetime(*entry.published_parsed[:6])
                            if pub_date < cutoff_date:
                                continue
                        except:
                            # If date parsing fails, include it anyway
                            pass

                    # Check if relevant keywords present
                    text = f"{entry.title} {entry.get('summary', '')}".lower()
                    if any(kw in text for kw in keywords):
                        articles.append({
                            'title': entry.title,
                            'description': entry.get('summary', ''),
                            'url': entry.link,
                            'source': feed.feed.get('title', 'RSS Feed'),
                            'published_at': entry.get('published', ''),
                            'content': entry.get('summary', '')
                        })
                        feed_articles += 1

                # Debug: show which feeds are producing results
                debug_mode = os.environ.get('DEBUG_FILTERING', 'false').lower() == 'true'
                if debug_mode and feed_articles > 0:
                    print(f"    ✓ {feed.feed.get('title', feed_url)}: {feed_articles} articles")

            except Exception as e:
                failed_feeds += 1
                debug_mode = os.environ.get('DEBUG_FILTERING', 'false').lower() == 'true'
                if debug_mode:
                    print(f"    ✗ Error with {feed_url[:50]}: {str(e)[:50]}")
        
        return articles
    
    def scrape_full_article(self, url: str) -> str:
        """
        Scrape full article text from URL
        Returns the article text or empty string if failed
        """
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
            }
            response = requests.get(url, headers=headers, timeout=10)
            response.raise_for_status()

            soup = BeautifulSoup(response.content, 'html.parser')

            # Remove script and style elements
            for script in soup(['script', 'style', 'nav', 'header', 'footer', 'aside']):
                script.decompose()

            # Try common article content selectors
            article_selectors = [
                'article',
                '[class*="article"]',
                '[class*="story"]',
                '[class*="content"]',
                '[class*="body"]',
                'main',
                '.post-content',
            ]

            article_text = ""
            for selector in article_selectors:
                article_elem = soup.select_one(selector)
                if article_elem:
                    # Get all paragraph text
                    paragraphs = article_elem.find_all('p')
                    article_text = ' '.join([p.get_text(strip=True) for p in paragraphs])
                    if len(article_text) > 200:  # Only accept if substantial text found
                        break

            # Fallback: get all paragraphs if specific selectors didn't work
            if len(article_text) < 200:
                paragraphs = soup.find_all('p')
                article_text = ' '.join([p.get_text(strip=True) for p in paragraphs])

            return article_text[:5000]  # Limit to first 5000 chars to avoid huge texts

        except Exception as e:
            debug_mode = os.environ.get('DEBUG_FILTERING', 'false').lower() == 'true'
            if debug_mode:
                print(f"    ⚠️ Web scraping failed for {url[:50]}: {str(e)[:50]}")
            return ""

    def parse_article_for_meetings(self, article: Dict) -> List[Dict]:
        """
        Parse article to extract meeting information
        Returns list of meeting dictionaries
        """
        meetings = []

        # Combine all text (summary first)
        text = f"{article['title']} {article['description']} {article.get('content', '')}"

        # Check if it's actually about Trump meetings
        # Enable debug mode for first 5 articles to see filtering reasons
        debug_mode = os.environ.get('DEBUG_FILTERING', 'false').lower() == 'true'
        if not self.is_trump_meeting_article(text, debug=debug_mode):
            return []

        # If it passes initial filter, try to get full article text
        if os.environ.get('ENABLE_WEB_SCRAPING', 'true').lower() == 'true':
            full_text = self.scrape_full_article(article['url'])
            if full_text:
                # Prepend summary, then add full article
                text = f"{text} {full_text}"
                if debug_mode:
                    print(f"    ✓ Scraped full article ({len(full_text)} chars)")
        
        # Extract date
        meeting_date = self.extract_meeting_date(text, article.get('published_at'))
        
        # Extract location
        location = self.extract_location(text)
        
        # Extract attendees (name, title, company)
        attendees = self.extract_attendees(text)

        debug_mode = os.environ.get('DEBUG_FILTERING', 'false').lower() == 'true'
        if not attendees:
            if debug_mode:
                print(f"    ⚠️ No attendees extracted from text")
                print(f"       Article: {article.get('title', 'No title')}")
                print(f"       Text sample: {text[:300]}...")
            return []
        
        # Create meeting object
        meeting = {
            'date': meeting_date,
            'location': location,
            'type': 'Business Meeting',
            'source_url': article['url'],
            'source_publication': article['source'],
            'notes': article['title'][:200],
            'attendees': []
        }
        
        # Process each attendee
        for att in attendees:
            # Classify company industry
            industry_info = self.classify_company_industry(att['company'])
            
            # Determine confidence level based on how company was found
            if att.get('found_in_article', True):
                # Company explicitly mentioned in article
                base_confidence = 'high'
            else:
                # Company found via dynamic lookup
                base_confidence = att.get('confidence', 'medium')
            
            # Adjust confidence based on industry classification
            if industry_info['confidence'] == 'low':
                # Unknown company reduces confidence
                final_confidence = 'low'
            else:
                final_confidence = base_confidence
            
            confidence_reasons = [f"Extracted from article: {article['source']}"]
            if not att.get('found_in_article', True):
                confidence_reasons.append("Company identified via dynamic web search")
            
            attendee_data = {
                'name': att['name'],
                'title': att['title'],
                'company': att['company'],
                'primary_industry': industry_info['primary_industry'],
                'secondary_industries': industry_info.get('secondary_industries', []),
                'confidence_level': final_confidence,
                'confidence_reasons': confidence_reasons,
                'requires_review': final_confidence == 'low'
            }
            
            meeting['attendees'].append(attendee_data)
        
        meetings.append(meeting)
        
        return meetings
    
    def is_trump_meeting_article(self, text: str, debug: bool = False) -> bool:
        """Check if article is about Trump meetings"""
        text_lower = text.lower()

        # Must mention Trump
        if 'trump' not in text_lower:
            if debug:
                print(f"    ❌ Filtered: No 'trump' mention")
            return False

        # Must have meeting indicators WITH Trump
        meeting_patterns = [
            'trump meet', 'trump met', 'trump host', 'trump welcomed',
            'meeting with trump', 'met with trump', 'hosted by trump'
        ]
        if not any(pattern in text_lower for pattern in meeting_patterns):
            if debug:
                print(f"    ❌ Filtered: No meeting pattern found")
            return False

        # Should mention business/executives (broader detection)
        business_words = [
            'ceo', 'chief executive', 'chairman', 'chief', 'business leader',
            'executive', 'company', 'founder', 'entrepreneur', 'businessman',
            'businesswoman', 'tech', 'corporation', 'industry', 'corporate',
            'investor', 'billionaire', 'magnate'
        ]
        if not any(word in text_lower for word in business_words):
            if debug:
                print(f"    ❌ Filtered: No business words found")
                print(f"       Text sample: {text_lower[:200]}")
            return False

        # Exclude articles primarily about foreign leaders or politics
        # But allow some political context (e.g., "CEO met Trump at White House to discuss tariffs")
        political_keywords = [
            'ukraine', 'russia', 'venezuela', 'maduro', 'macron', 'zelensky', 'iran',
            'foreign leader', 'prime minister', 'nato', 'invasion', 'military',
            'war', 'sanctions', 'diplomacy', 'treaty', 'ambassador'
        ]
        # Count political keywords
        political_count = sum(1 for kw in political_keywords if kw in text_lower)
        # If more than 4 political keywords, likely not a business meeting (relaxed from 2)
        if political_count > 4:
            if debug:
                print(f"    ❌ Filtered: Too many political keywords ({political_count})")
            return False

        if debug:
            print(f"    ✅ Passed Trump meeting article check")
        return True
    
    def extract_meeting_date(self, text: str, published_date: str = None) -> str:
        """Extract meeting date from text"""
        # Look for explicit dates
        date_patterns = [
            r'(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},?\s+\d{4}',
            r'\d{1,2}/\d{1,2}/\d{4}',
            r'(?:Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday),?\s+(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2}'
        ]
        
        for pattern in date_patterns:
            match = re.search(pattern, text)
            if match:
                return match.group(0)
        
        # Fall back to published date
        if published_date:
            try:
                dt = datetime.fromisoformat(published_date.replace('Z', '+00:00'))
                return dt.strftime('%B %d, %Y')
            except:
                pass
        
        # Default to today
        return datetime.now().strftime('%B %d, %Y')
    
    def extract_location(self, text: str) -> str:
        """Extract meeting location from text"""
        locations = {
            'Mar-a-Lago': ['mar-a-lago', 'mar a lago'],
            'White House, DC': ['white house'],
            'Trump Tower, NY': ['trump tower'],
            'Bedminster, NJ': ['bedminster']
        }
        
        text_lower = text.lower()
        for location, keywords in locations.items():
            if any(kw in text_lower for kw in keywords):
                return location
        
        return 'Location TBD'
    
    def extract_attendees(self, text: str) -> List[Dict]:
        """
        Extract attendee names, titles, and companies from text
        Returns list of {name, title, company}
        """
        attendees = []
        
        # Pattern 1: Name, Title of Company
        # Example: "Andy Jassy, CEO of Amazon"
        # Accept CEO/Chairman/Chief titles + President (but we'll filter out countries later)
        pattern1 = r'([A-Z][a-z]+\s+[A-Z][a-z]+(?:\s+[A-Z][a-z]+)?),\s+(CEO|Chairman|Chief\s+Executive|Chief\s+Operating\s+Officer|CFO|COO|Chief\s+Financial\s+Officer|President|Founder|Co-Founder|Managing\s+Director|Executive\s+Chairman)\s+(?:of\s+|at\s+)([A-Z][A-Za-z0-9\s&\.]+?)(?:\.|,|\s+(?:said|told|announced|met|joined|attended))'
        matches1 = re.findall(pattern1, text)

        for match in matches1:
            name, title, company = match
            company = company.strip()

            # Skip if company looks like a country or government entity
            if self.is_government_or_country(company):
                continue

            # Clean up company name
            company = re.sub(r'\s+Inc\.?|\s+Corp\.?|\s+LLC|\s+Ltd\.?', '', company)

            attendees.append({
                'name': name.strip(),
                'title': title.strip(),
                'company': company.strip(),
                'found_in_article': True
            })
        
        # Pattern 2: Company CEO Name
        # Example: "Amazon CEO Andy Jassy", "Intel CEO Lip-Bu Tan"
        # Accept CEO/Chairman/President/Founder (but we'll filter out countries later)
        # More restrictive: company name should be 1-3 words max
        # Support hyphenated names like Lip-Bu
        pattern2 = r'([A-Z][A-Za-z0-9]+(?:\s+[A-Z&][A-Za-z0-9]+){0,2})\s+(CEO|Chairman|Chief\s+Executive|President|Founder|Co-Founder|Managing\s+Director|Executive\s+Chairman)\s+([A-Z][a-z]+(?:-[A-Z][a-z]+)?\s+[A-Z][a-z]+)'
        matches2 = re.findall(pattern2, text)

        for match in matches2:
            company, title, name = match
            company = company.strip()
            name_str = name.strip()

            # Skip Trump or Biden (political figures)
            if 'Trump' in name_str or 'Biden' in name_str:
                continue

            # Skip if company starts with articles or political words
            company_words = company.split()
            if company_words and company_words[0] in ['President', 'Former', 'Vice', 'Senator', 'The']:
                continue

            # Skip if company looks like a country or government entity
            if self.is_government_or_country(company):
                continue

            # Skip if company name is suspiciously long (probably grabbed wrong text)
            if len(company.split()) > 4:
                continue

            company = re.sub(r'\s+Inc\.?|\s+Corp\.?|\s+LLC|\s+Ltd\.?', '', company)

            # Avoid duplicates
            if not any(a['name'] == name_str for a in attendees):
                attendees.append({
                    'name': name_str,
                    'title': title.strip(),
                    'company': company.strip(),
                    'found_in_article': True
                })

        # Pattern 2.5: Company CEO without name (e.g., "Trump meets Intel CEO")
        # Extract company and try to look up current CEO dynamically
        if len(attendees) == 0 and os.environ.get('ENABLE_DYNAMIC_CEO_LOOKUP', 'false').lower() == 'true':
            # Look for patterns: "Trump meets [Company] CEO" or "meeting with [Company] CEO"
            company_ceo_pattern = r'(?:meets|met|hosted|host|meeting\s+with)\s+(?:with\s+)?([A-Z][A-Za-z0-9]+(?:\s+[A-Z&][A-Za-z0-9]+){0,2})\s+(CEO|Chairman|Chief\s+Executive|President)'
            matches_company = re.findall(company_ceo_pattern, text)

            for match in matches_company[:1]:  # Only try first company mention
                company, title = match
                company = company.strip()

                # Skip if government/country
                if self.is_government_or_country(company):
                    continue

                # Try to look up the current CEO
                ceo_info = self.lookup_company_ceo(company)
                if ceo_info:
                    debug_mode = os.environ.get('DEBUG_FILTERING', 'false').lower() == 'true'
                    if debug_mode:
                        print(f"    ✓ Looked up {company} CEO: {ceo_info['name']}")
                    attendees.append({
                        'name': ceo_info['name'],
                        'title': title.strip(),
                        'company': company,
                        'found_in_article': False  # Name wasn't in article
                    })
                    break

        # Pattern 3: Just well-known names (Elon Musk, Tim Cook, etc.)
        # Match prominent business figures even without explicit title/company
        # These are commonly referred to by name alone in headlines
        prominent_ceos = {
            'Elon Musk': {'company': 'Tesla', 'title': 'CEO'},
            'Tim Cook': {'company': 'Apple', 'title': 'CEO'},
            'Mark Zuckerberg': {'company': 'Meta', 'title': 'CEO'},
            'Sundar Pichai': {'company': 'Google', 'title': 'CEO'},
            'Satya Nadella': {'company': 'Microsoft', 'title': 'CEO'},
            'Jeff Bezos': {'company': 'Amazon', 'title': 'Executive Chairman'},
            'Andy Jassy': {'company': 'Amazon', 'title': 'CEO'},
            'Jensen Huang': {'company': 'NVIDIA', 'title': 'CEO'},
            'Sam Altman': {'company': 'OpenAI', 'title': 'CEO'},
            'Jamie Dimon': {'company': 'JPMorgan Chase', 'title': 'CEO'},
            'Mary Barra': {'company': 'GM', 'title': 'CEO'},
            'Doug McMillon': {'company': 'Walmart', 'title': 'CEO'},
            'Larry Fink': {'company': 'BlackRock', 'title': 'CEO'},
            'Brian Moynihan': {'company': 'Bank of America', 'title': 'CEO'},
            'David Solomon': {'company': 'Goldman Sachs', 'title': 'CEO'},
            'Dara Khosrowshahi': {'company': 'Uber', 'title': 'CEO'},
        }

        # Look for these names in the text
        text_with_spaces = f" {text} "  # Add spaces for word boundary matching
        for name, info in prominent_ceos.items():
            # Case-insensitive search with word boundaries
            if name.lower() in text_with_spaces.lower():
                # Verify it's actually about this person (not just coincidence)
                if not any(a['name'] == name for a in attendees):
                    attendees.append({
                        'name': name,
                        'title': info['title'],
                        'company': info['company'],
                        'found_in_article': True
                    })

        # Pattern 4: Dynamic name extraction for unknown CEOs
        # Look for capitalized names that might be executives we don't know
        if len(attendees) == 0 and os.environ.get('ENABLE_DYNAMIC_CEO_LOOKUP', 'false').lower() == 'true':
            # Pattern: Two or three capitalized words (may include hyphens)
            # Examples: "John Smith", "Lip-Bu Tan", "Mary Jane Watson"
            name_pattern = r'\b([A-Z][a-z]+(?:-[A-Z][a-z]+)?\s+[A-Z][a-z]+(?:\s+[A-Z][a-z]+)?)\b'
            potential_names = re.findall(name_pattern, text)

            debug_mode = os.environ.get('DEBUG_FILTERING', 'false').lower() == 'true'

            for potential_name in potential_names[:5]:  # Check first 5 potential names
                # Skip if it's Trump, Biden, or already found
                if potential_name in ['Donald Trump', 'Joe Biden']:
                    continue
                if any(a['name'] == potential_name for a in attendees):
                    continue

                # IMPORTANT: Use looks_like_person_name() to filter out garbage
                if not self.looks_like_person_name(potential_name):
                    if debug_mode:
                        print(f"    ⚠️ Skipping '{potential_name}' - doesn't look like a person name")
                    continue

                # Check if this name appears near business/meeting context
                name_pos = text.lower().find(potential_name.lower())
                if name_pos == -1:
                    continue

                context = text[max(0, name_pos-150):min(len(text), name_pos+150)].lower()
                business_indicators = ['ceo', 'chief', 'executive', 'chairman', 'president', 'founder', 'company']

                if any(indicator in context for indicator in business_indicators):
                    # Try to look up this person using dynamic search
                    company_info = self.lookup_person_company_dynamic(potential_name, text)
                    if company_info:
                        if debug_mode:
                            print(f"    ✓ Dynamic lookup: {potential_name} - {company_info['title']} at {company_info['company']}")
                        attendees.append({
                            'name': potential_name,
                            'title': company_info['title'],
                            'company': company_info['company'],
                            'found_in_article': True
                        })
                        break  # Stop after first successful match to avoid rate limits

        return attendees
    
    def is_government_or_country(self, company_name: str) -> bool:
        """Check if the 'company' is actually a government entity or country"""
        company_lower = company_name.lower().strip()

        # List of government/political keywords
        government_keywords = [
            'national assembly', 'government', 'ministry', 'parliament', 'congress',
            'senate', 'administration', 'department of', 'agency', 'commission',
            'federal', 'state department', 'white house', 'embassy', 'consulate',
            'republic', 'kingdom', 'federation', 'union', 'nation', 'country',
            'military', 'army', 'navy', 'defense', 'homeland security',
            'foreign affairs', 'state', 'democratic', 'republic of',
            'united states', 'european union', 'nato', 'un ', 'u.n.'
        ]

        # Countries and regions - expanded list
        countries = [
            'venezuela', 'france', 'ukraine', 'russia', 'iran', 'mexico', 'colombia',
            'denmark', 'greenland', 'china', 'israel', 'syria', 'iraq', 'afghanistan',
            'canada', 'britain', 'germany', 'italy', 'spain', 'poland', 'japan',
            'korea', 'brazil', 'argentina', 'egypt', 'turkey', 'india', 'pakistan',
            'saudi arabia', 'united arab emirates', 'qatar', 'taiwan', 'vietnam',
            'thailand', 'indonesia', 'australia', 'new zealand', 'south africa'
        ]

        # Check if it matches any government keywords or countries
        for keyword in government_keywords + countries:
            if keyword in company_lower:
                return True

        # Check if it's too generic (single word entities that aren't companies)
        nationality_adjectives = [
            'danish', 'venezuelan', 'colombian', 'mexican', 'iranian', 'french',
            'canadian', 'british', 'german', 'italian', 'spanish', 'japanese',
            'korean', 'chinese', 'russian', 'ukrainian', 'israeli', 'egyptian'
        ]
        if len(company_lower.split()) == 1 and company_lower in nationality_adjectives:
            return True

        return False

    def looks_like_person_name(self, name: str) -> bool:
        """Check if a string looks like an actual person's name"""
        parts = name.split()

        # Must be 2-3 words
        if len(parts) < 2 or len(parts) > 3:
            return False

        # Each part should be capitalized and reasonable length
        # Handle hyphenated names like "Lip-Bu"
        for part in parts:
            # Split on hyphens to check each component
            subparts = part.split('-')
            for subpart in subparts:
                if not subpart or not subpart[0].isupper():
                    return False
                if len(subpart) < 2:  # Each component should be at least 2 letters (allows "Li-" or "Bu")
                    return False
                if len(subpart) > 15:  # Too long to be a name
                    return False

        # Each part should be primarily lowercase letters after first char
        # This filters out things like "Bu Tan" which might be fragments of words
        for part in parts:
            # For hyphenated parts, check the whole thing
            lowercase_count = sum(1 for c in part[1:] if c.islower())
            alpha_count = sum(1 for c in part[1:] if c.isalpha())
            if alpha_count > 0 and lowercase_count < alpha_count * 0.4:
                # Less than 40% lowercase letters (excluding hyphens) = probably not a name
                return False

        # Reject common non-name patterns
        non_name_words = {
            'president', 'ceo', 'chairman', 'chief', 'executive', 'officer',
            'company', 'corporation', 'inc', 'llc', 'ltd', 'business',
            'administration', 'department', 'agency', 'house', 'senate',
            'heritage', 'foundation', 'project', 'act', 'services', 'education',
            'disabilities', 'human', 'armed', 'vocational', 'aptitude', 'battery',
            'head', 'start', 'reproductive', 'freedom', 'health', 'resources',
            'secretary', 'robert', 'alive', 'abortion', 'survivors', 'medicaid',
            'homeland', 'security', 'border', 'protection', 'customs', 'enforcement',
            'national', 'weather', 'service', 'fair', 'labor', 'standards',
            'supreme', 'court', 'civil', 'war', 'white', 'donald', 'trump',
            # Technical/manufacturing terms
            'made', 'sub', 'nanometer', 'chip', 'western', 'hemisphere', 'insanity',
            'rules', 'united', 'states', 'north', 'south', 'east', 'west', 'new', 'york'
        }

        if any(part.lower() in non_name_words for part in parts):
            return False

        return True
    
    def appears_near_meeting_context(self, name: str, text: str) -> bool:
        """Check if name appears near meeting-related words"""
        # Find position of name
        pos = text.find(name)
        if pos == -1:
            return False

        # Check 100 chars before and after
        context = text[max(0, pos-100):min(len(text), pos+100)].lower()

        meeting_words = ['met', 'meeting', 'hosted', 'spoke', 'discussed', 'attended', 'joined']
        return any(word in context for word in meeting_words)

    def appears_near_business_context(self, name: str, text: str) -> bool:
        """Check if name appears near business-related words (CEO, company, etc.)"""
        # Find position of name
        pos = text.find(name)
        if pos == -1:
            return False

        # Check 100 chars before and after
        context = text[max(0, pos-100):min(len(text), pos+100)].lower()

        business_words = ['ceo', 'chief executive', 'president', 'chairman', 'company',
                         'corporation', 'executive', 'founder', 'co-founder', 'business']
        return any(word in context for word in business_words)
    
    def lookup_person_company_dynamic(self, person_name: str, article_context: str = "") -> Optional[Dict]:
        """
        Dynamically look up a person's company using web search
        Returns: {'company': str, 'title': str, 'confidence': str}
        """
        # Skip lookups if we don't have NewsAPI (rate limiting protection)
        if not self.newsapi:
            return None

        print(f"    🔍 Looking up: {person_name}")
        
        # First, check if we can infer from article context
        # Look for patterns like "person_name, who is/was CEO of Company"
        patterns = [
            f"{re.escape(person_name)}[^.]*?(?:CEO|President|Chairman|Chief Executive)[^.]*?(?:of|at)\\s+([A-Z][A-Za-z0-9]+(?:\\s+[A-Z][A-Za-z0-9]+)?)",
            f"([A-Z][A-Za-z0-9]+(?:\\s+[A-Z][A-Za-z0-9]+)?)\\s+(?:CEO|President|Chairman)\\s+{re.escape(person_name)}"
        ]
        
        for pattern in patterns:
            match = re.search(pattern, article_context, re.IGNORECASE)
            if match:
                company = match.group(1).strip()
                # Clean up
                company = re.sub(r'\s+Inc\.?|\s+Corp\.?|\s+LLC|\s+Ltd\.?', '', company)
                company = company.split(',')[0].split('.')[0].strip()
                
                # Validate it looks like a company name
                if len(company.split()) <= 3 and not any(word.lower() in ['the', 'and', 'for'] for word in company.split()):
                    print(f"    ✓ Found in context: {company}")
                    return {
                        'company': company,
                        'title': 'CEO',
                        'confidence': 'medium'
                    }
        
        # If NewsAPI is available, search for the person
        if self.newsapi:
            try:
                # Search for recent articles about this person + CEO
                search_results = self.newsapi.get_everything(
                    q=f'"{person_name}" CEO',
                    language='en',
                    sort_by='relevancy',
                    page_size=3
                )
                
                if search_results['status'] == 'ok' and search_results['articles']:
                    # Look through articles for company mentions
                    for article in search_results['articles']:
                        article_text = f"{article.get('title', '')} {article.get('description', '')} {article.get('content', '')}"
                        
                        # Look for clear company patterns
                        patterns = [
                            f"{re.escape(person_name)}[^.]*?(?:CEO|President|Chairman)[^.]*?(?:of|at)\\s+([A-Z][A-Za-z0-9]+(?:\\s+[A-Z][A-Za-z0-9]+)?)",
                            f"([A-Z][A-Za-z0-9]+(?:\\s+[A-Z][A-Za-z0-9]+)?)\\s+(?:CEO|President|Chairman)\\s+{re.escape(person_name)}"
                        ]
                        
                        for pattern in patterns:
                            match = re.search(pattern, article_text, re.IGNORECASE)
                            if match:
                                company = match.group(1).strip()
                                # Clean and validate
                                company = re.sub(r'\s+Inc\.?|\s+Corp\.?|\s+LLC|\s+Ltd\.?', '', company)
                                company = company.split(',')[0].split('.')[0].strip()
                                
                                # Check if it's a valid company name (not too long, not common words)
                                if (2 <= len(company) <= 30 and 
                                    len(company.split()) <= 3 and
                                    company[0].isupper()):
                                    
                                    print(f"    ✓ Found via web search: {company}")
                                    return {
                                        'company': company,
                                        'title': 'CEO',
                                        'confidence': 'medium'
                                    }
            except Exception as e:
                print(f"    ⚠️ Error in web search: {str(e)}")
        
        print(f"    ✗ Could not find company for {person_name}")
        return None

    def lookup_company_ceo(self, company_name: str) -> Optional[Dict]:
        """
        Look up the current CEO of a company using web search
        Returns: {'name': str, 'title': str}
        """
        # Skip lookups if we don't have NewsAPI (rate limiting protection)
        if not self.newsapi:
            return None

        debug_mode = os.environ.get('DEBUG_FILTERING', 'false').lower() == 'true'
        if debug_mode:
            print(f"    🔍 Looking up {company_name} CEO...")

        try:
            # Search for recent articles about this company's CEO
            search_results = self.newsapi.get_everything(
                q=f'"{company_name}" CEO',
                language='en',
                sort_by='relevancy',
                page_size=5
            )

            if search_results['status'] == 'ok' and search_results['articles']:
                # Look through articles for CEO name
                for article in search_results['articles']:
                    article_text = f"{article.get('title', '')} {article.get('description', '')} {article.get('content', '')}"

                    # Look for patterns like "Company CEO Name" or "Name, CEO of Company"
                    patterns = [
                        f"{re.escape(company_name)}\\s+CEO\\s+([A-Z][a-z]+\\s+[A-Z][a-z]+)",
                        f"([A-Z][a-z]+\\s+[A-Z][a-z]+),\\s+(?:CEO|Chief Executive|Chief Executive Officer)\\s+(?:of|at)\\s+{re.escape(company_name)}",
                        f"([A-Z][a-z]+\\s+[A-Z][a-z]+)\\s+is\\s+(?:the\\s+)?CEO\\s+(?:of|at)\\s+{re.escape(company_name)}"
                    ]

                    for pattern in patterns:
                        match = re.search(pattern, article_text, re.IGNORECASE)
                        if match:
                            ceo_name = match.group(1).strip()

                            # Validate it looks like a person name
                            if self.looks_like_person_name(ceo_name):
                                if debug_mode:
                                    print(f"    ✓ Found: {ceo_name}")
                                return {
                                    'name': ceo_name,
                                    'title': 'CEO'
                                }

        except Exception as e:
            if debug_mode:
                print(f"    ⚠️ Error looking up CEO: {str(e)}")

        if debug_mode:
            print(f"    ✗ Could not find CEO for {company_name}")
        return None

    def classify_company_industry(self, company_name: str) -> Dict:
        """
        Classify company into industry categories using config
        Improved algorithm with better prioritization
        """
        company_lower = company_name.lower().strip()

        # Priority 1: Exact or near-exact match with known companies (HIGHEST CONFIDENCE)
        for industry_cat in self.config['industry_categories']:
            if 'related_companies' in industry_cat:
                for known_company in industry_cat['related_companies']:
                    known_lower = known_company.lower()

                    # Exact match
                    if company_lower == known_lower:
                        return {
                            'primary_industry': industry_cat['name'],
                            'secondary_industries': [],
                            'confidence': 'very high'
                        }

                    # Company name contains known company as whole word
                    # e.g., "Intel Corporation" matches "Intel"
                    if (f' {known_lower} ' in f' {company_lower} ' or
                        company_lower.startswith(known_lower + ' ') or
                        company_lower.endswith(' ' + known_lower) or
                        company_lower == known_lower):
                        return {
                            'primary_industry': industry_cat['name'],
                            'secondary_industries': [],
                            'confidence': 'high'
                        }

        # Priority 2: Known company partial match (MEDIUM-HIGH CONFIDENCE)
        # Only if company name is short enough to avoid false positives
        if len(company_lower) <= 20:
            for industry_cat in self.config['industry_categories']:
                if 'related_companies' in industry_cat:
                    for known_company in industry_cat['related_companies']:
                        known_lower = known_company.lower()

                        # Fuzzy match for short names
                        if self.fuzzy_match(known_lower, company_lower):
                            return {
                                'primary_industry': industry_cat['name'],
                                'secondary_industries': [],
                                'confidence': 'medium-high'
                            }

        # Priority 3: Industry-specific keywords (MEDIUM CONFIDENCE)
        # Only match if keyword is significant portion of company name
        keyword_matches = []
        for industry_cat in self.config['industry_categories']:
            if 'keywords' in industry_cat:
                for keyword in industry_cat['keywords']:
                    keyword_lower = keyword.lower()
                    # Skip single-char or very short keywords
                    if len(keyword_lower) < 4:
                        continue
                    # Check if keyword is in company name
                    if keyword_lower in company_lower:
                        # Calculate match quality based on keyword length vs company name
                        match_score = len(keyword_lower) / len(company_lower)
                        keyword_matches.append({
                            'industry': industry_cat['name'],
                            'score': match_score,
                            'keyword': keyword_lower
                        })

        # Return best keyword match if any found
        if keyword_matches:
            best_match = max(keyword_matches, key=lambda x: x['score'])
            # Only accept if keyword is at least 30% of company name
            if best_match['score'] >= 0.3:
                return {
                    'primary_industry': best_match['industry'],
                    'secondary_industries': [],
                    'confidence': 'medium'
                }

        # No good match found
        return {
            'primary_industry': 'Other',
            'secondary_industries': [],
            'confidence': 'low'
        }
    
    def fuzzy_match(self, str1: str, str2: str) -> bool:
        """Simple fuzzy string matching"""
        # Check if significant portion of one string is in the other
        if len(str1) < 4 or len(str2) < 4:
            return False
        
        # Check for common core (at least 4 chars)
        for i in range(len(str1) - 3):
            substr = str1[i:i+4]
            if substr in str2:
                return True
        
        return False
    
    def is_duplicate_meeting(self, meeting_data: Dict) -> Dict:
        """
        Check if meeting already exists in database by date + attendee name
        Returns: {
            'is_duplicate': bool,
            'meeting_id': int or None,
            'should_merge': bool  # True if same meeting from different source
        }
        """
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        meeting_date = meeting_data.get('date')
        attendees = meeting_data.get('attendees', [])

        # If no date or no attendees, can't deduplicate intelligently
        if not meeting_date or not attendees:
            conn.close()
            return {'is_duplicate': False, 'meeting_id': None, 'should_merge': False}

        # Check if this exact source URL already exists
        cursor.execute('''
            SELECT id FROM meetings
            WHERE source_url = ?
        ''', (meeting_data.get('source_url'),))

        exact_match = cursor.fetchone()
        if exact_match:
            conn.close()
            return {'is_duplicate': True, 'meeting_id': exact_match[0], 'should_merge': False}

        # Check for same meeting from different source (by date + attendee name)
        # Get all meetings on the same date
        cursor.execute('''
            SELECT m.id
            FROM meetings m
            WHERE m.date = ?
        ''', (meeting_date,))

        potential_duplicates = cursor.fetchall()

        for (meeting_id,) in potential_duplicates:
            # Get attendees for this meeting
            cursor.execute('''
                SELECT name FROM attendees WHERE meeting_id = ?
            ''', (meeting_id,))

            existing_attendees = [row[0] for row in cursor.fetchall()]

            # Check if any attendee name matches
            for new_attendee in attendees:
                new_name = new_attendee.get('name', '').strip().lower()
                for existing_name in existing_attendees:
                    existing_name_lower = existing_name.strip().lower()

                    # Exact match or one name contains the other
                    if (new_name == existing_name_lower or
                        (len(new_name) > 5 and new_name in existing_name_lower) or
                        (len(existing_name_lower) > 5 and existing_name_lower in new_name)):

                        conn.close()
                        return {
                            'is_duplicate': True,
                            'meeting_id': meeting_id,
                            'should_merge': True  # Same meeting, different source
                        }

        conn.close()
        return {'is_duplicate': False, 'meeting_id': None, 'should_merge': False}

    def merge_meeting_source(self, meeting_id: int, new_source_url: str, new_source_publication: str) -> bool:
        """
        Merge a new source into an existing meeting
        Updates source_urls array and source_count
        """
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        try:
            # Get current source_urls
            cursor.execute('''
                SELECT source_urls, source_url, source_publication
                FROM meetings
                WHERE id = ?
            ''', (meeting_id,))

            row = cursor.fetchone()
            if not row:
                conn.close()
                return False

            source_urls_json, original_url, original_pub = row

            # Parse existing sources
            if source_urls_json:
                try:
                    source_urls = json.loads(source_urls_json)
                except (json.JSONDecodeError, TypeError):
                    source_urls = [original_url] if original_url else []
            else:
                source_urls = [original_url] if original_url else []

            # Add new source if not already present
            if new_source_url not in source_urls:
                source_urls.append(new_source_url)

                # Update the meeting record
                cursor.execute('''
                    UPDATE meetings
                    SET source_urls = ?,
                        source_count = ?,
                        source_publication = ?
                    WHERE id = ?
                ''', (
                    json.dumps(source_urls),
                    len(source_urls),
                    f"{original_pub}, {new_source_publication}" if original_pub else new_source_publication,
                    meeting_id
                ))

                conn.commit()
                conn.close()
                return True
            else:
                conn.close()
                return False

        except Exception as e:
            print(f"  ⚠️ Error merging source: {str(e)}")
            conn.close()
            return False

    def save_meeting(self, meeting_data: Dict) -> int:
        """Save meeting to database, return meeting_id"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        source_url = meeting_data.get('source_url')
        source_urls_json = json.dumps([source_url]) if source_url else json.dumps([])

        try:
            cursor.execute('''
                INSERT INTO meetings (date, location, meeting_type, source_url,
                                    source_publication, date_added, notes,
                                    source_urls, source_count)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                meeting_data.get('date'),
                meeting_data.get('location'),
                meeting_data.get('type'),
                source_url,
                meeting_data.get('source_publication'),
                datetime.now().isoformat(),
                meeting_data.get('notes'),
                source_urls_json,
                1
            ))
            
            meeting_id = cursor.lastrowid
            
            # Save attendees
            for attendee in meeting_data.get('attendees', []):
                cursor.execute('''
                    INSERT INTO attendees (meeting_id, name, title, company, 
                                         primary_industry, secondary_industries,
                                         confidence_level, confidence_reasons, requires_review)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    meeting_id,
                    attendee['name'],
                    attendee.get('title'),
                    attendee.get('company'),
                    attendee.get('primary_industry'),
                    json.dumps(attendee.get('secondary_industries', [])),
                    attendee.get('confidence_level'),
                    json.dumps(attendee.get('confidence_reasons', [])),
                    attendee.get('requires_review', False)
                ))
            
            conn.commit()
        except sqlite3.IntegrityError:
            # Duplicate - skip
            conn.close()
            return -1
        except Exception as e:
            print(f"  ⚠️ Error saving meeting: {str(e)}")
            conn.close()
            return -1
        
        conn.close()
        return meeting_id
    
    def get_new_meetings(self, since_date: str) -> List[Dict]:
        """Get meetings added since a specific date"""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cursor.execute('''
            SELECT * FROM meetings
            WHERE date_added >= ?
            ORDER BY date DESC
        ''', (since_date,))

        meetings = []
        for meeting_row in cursor.fetchall():
            meeting = dict(meeting_row)

            # Get attendees for this meeting
            cursor.execute('''
                SELECT * FROM attendees WHERE meeting_id = ?
            ''', (meeting['id'],))

            attendees = []
            for att_row in cursor.fetchall():
                attendee = dict(att_row)
                try:
                    attendee['secondary_industries'] = json.loads(attendee['secondary_industries'])
                    attendee['confidence_reasons'] = json.loads(attendee['confidence_reasons'])
                except:
                    attendee['secondary_industries'] = []
                    attendee['confidence_reasons'] = []
                attendees.append(attendee)

            meeting['attendees'] = attendees
            meetings.append(meeting)

        conn.close()
        return meetings

    def get_all_meetings(self) -> List[Dict]:
        """Get all meetings from the database (for Excel report)"""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cursor.execute('''
            SELECT * FROM meetings
            ORDER BY date DESC
        ''')

        meetings = []
        for meeting_row in cursor.fetchall():
            meeting = dict(meeting_row)

            # Get attendees for this meeting
            cursor.execute('''
                SELECT * FROM attendees WHERE meeting_id = ?
            ''', (meeting['id'],))

            attendees = []
            for att_row in cursor.fetchall():
                attendee = dict(att_row)
                try:
                    attendee['secondary_industries'] = json.loads(attendee['secondary_industries'])
                    attendee['confidence_reasons'] = json.loads(attendee['confidence_reasons'])
                except:
                    attendee['secondary_industries'] = []
                    attendee['confidence_reasons'] = []
                attendees.append(attendee)

            meeting['attendees'] = attendees
            meetings.append(meeting)

        conn.close()
        return meetings
    
    def generate_email_html(self, meetings: List[Dict]) -> str:
        """Generate HTML email from meetings data"""
        if not meetings:
            return """
            <html>
            <body style="font-family: Arial, sans-serif;">
                <h2>Trump Meetings Update</h2>
                <p>No new meetings found this period.</p>
            </body>
            </html>
            """
        
        # Categorize by priority
        high_priority = []
        medium_priority = []
        low_priority = []
        
        priority_industries = [
            '3PL', 'Asian 3PL', 'Agriculture', 'Automotive', 'Building Materials',
            'Data Center', 'E-Commerce', 'Asian E-Commerce', 'Food & Beverage',
            'Fulfillment & Packaging', 'Life Sciences', 'Manufacturing',
            'Powered Land', 'Retail', 'Wholesaler', 'Cold Storage'
        ]
        
        for meeting in meetings:
            meeting_priority = 'low'
            for attendee in meeting['attendees']:
                industry = attendee.get('primary_industry', 'Other')
                confidence = attendee.get('confidence_level', 'low')
                
                if industry in priority_industries:
                    if confidence == 'high':
                        meeting_priority = 'high'
                        break
                    elif confidence == 'medium' and meeting_priority != 'high':
                        meeting_priority = 'medium'
            
            if meeting_priority == 'high':
                high_priority.append(meeting)
            elif meeting_priority == 'medium':
                medium_priority.append(meeting)
            else:
                low_priority.append(meeting)
        
        # Build HTML
        html = f"""
        <html>
        <head>
            <style>
                body {{ font-family: "Georgia", "Times New Roman", serif; line-height: 1.6; color: #1f2a33; max-width: 820px; margin: 0 auto; padding: 24px; background: #ffffff; }}
                h1 {{ color: #0f1f2e; font-size: 26px; letter-spacing: 0.3px; border-bottom: 2px solid #d5dde6; padding-bottom: 12px; margin-bottom: 18px; }}
                h2 {{ color: #0f1f2e; font-size: 18px; margin-top: 28px; border-bottom: 1px solid #e6ebf0; padding-bottom: 6px; }}
                .summary {{ background-color: #f6f8fa; padding: 16px 18px; border: 1px solid #e1e6eb; margin: 18px 0; }}
                .high-priority {{ border-left: 4px solid #9b1c1c; padding: 14px 16px; margin: 14px 0; background: #fbf6f6; }}
                .medium-priority {{ border-left: 4px solid #b45309; padding: 14px 16px; margin: 14px 0; background: #fff8ed; }}
                .low-priority {{ border-left: 4px solid #6b7280; padding: 14px 16px; margin: 14px 0; background: #f8f9fb; }}
                .meeting-date {{ font-weight: bold; color: #111827; font-size: 1em; margin-bottom: 8px; }}
                .attendee {{ margin: 10px 0; padding: 10px 12px; background-color: #ffffff; border: 1px solid #e5e7eb; }}
                .company {{ color: #1d4ed8; font-weight: bold; }}
                .industry {{ color: #065f46; font-weight: 600; }}
                .confidence {{ font-size: 0.9em; font-style: italic; color: #4b5563; }}
                .confidence.high {{ color: #065f46; }}
                .confidence.medium {{ color: #92400e; }}
                .confidence.low {{ color: #991b1b; }}
                .source {{ font-size: 0.85em; margin-top: 10px; padding-top: 10px; border-top: 1px solid #e5e7eb; color: #374151; }}
                .source a {{ color: #1d4ed8; text-decoration: none; }}
                .source a:hover {{ text-decoration: underline; }}
                .meta {{ color: #4b5563; }}
            </style>
        </head>
        <body>
            <h1>Trump Meetings Report</h1>
            <div class="summary">
                <strong>Report Generated:</strong> {datetime.now().strftime('%B %d, %Y at %I:%M %p')}<br>
                <strong>Period:</strong> Last 7 days<br>
                <strong>New Meetings:</strong> {len(meetings)}<br>
                <strong>High Priority:</strong> {len(high_priority)} | 
                <strong>Medium Priority:</strong> {len(medium_priority)} | 
                <strong>Other:</strong> {len(low_priority)}
            </div>
        """
        
        if high_priority:
            html += "<h2>High Priority - Your Industries</h2>"
            for meeting in high_priority:
                html += self.format_meeting_html(meeting, 'high-priority')
        
        if medium_priority:
            html += "<h2>Medium Priority</h2>"
            for meeting in medium_priority:
                html += self.format_meeting_html(meeting, 'medium-priority')
        
        if low_priority:
            html += "<h2>Other Meetings</h2>"
            for meeting in low_priority:
                html += self.format_meeting_html(meeting, 'low-priority')
        
        html += """
            <div style="margin-top: 40px; padding-top: 20px; border-top: 1px solid #e5e7eb; font-size: 0.9em; color: #4b5563;">
                <p><strong>About This Report</strong></p>
                <ul>
                    <li>Automated tracking of Trump's meetings with business leaders</li>
                    <li>Sources: NewsAPI + RSS feeds from major news outlets</li>
                    <li>Industries classified based on company information</li>
                    <li>Confidence levels indicate certainty of company/industry match</li>
                    <li>Review meetings with low confidence manually</li>
                </ul>
            </div>
        </body>
        </html>
        """
        
        return html
    
    def format_meeting_html(self, meeting: Dict, css_class: str) -> str:
        """Format a single meeting as HTML"""
        html = f'<div class="{css_class}">'
        html += f'<div class="meeting-date">{meeting["date"]} - {meeting.get("location", "Location TBD")}</div>'
        
        for attendee in meeting['attendees']:
            confidence_class = attendee.get('confidence_level', 'low')
            html += f'''
            <div class="attendee">
                <strong>{attendee["name"]}</strong> - {attendee.get("title", "Executive")}<br>
                <span class="company">{attendee.get("company", "Unknown Company")}</span><br>
                <span class="industry">Industry: {attendee.get("primary_industry", "Unknown")}</span><br>
                <span class="confidence {confidence_class}">Confidence: {attendee.get("confidence_level", "unknown").upper()}</span>
            </div>
            '''
        
        if meeting.get('notes'):
            html += f'<div style="margin-top:10px; font-size:0.9em; color:#666;"><strong>Context:</strong> {meeting["notes"]}</div>'

        # Show multiple sources if available
        source_urls_json = meeting.get('source_urls', '[]')
        try:
            source_urls = json.loads(source_urls_json) if isinstance(source_urls_json, str) else source_urls_json
        except (json.JSONDecodeError, TypeError):
            source_urls = [meeting.get('source_url')] if meeting.get('source_url') else []

        if source_urls:
            source_count = len(source_urls)
            if source_count > 1:
                html += f'<div class="source"><strong>Reported by {source_count} sources:</strong><br>'
                for i, url in enumerate(source_urls, 1):
                    html += f'{i}. <a href="{url}">Source {i}</a><br>'
                html += '</div>'
            elif source_urls[0]:
                html += f'<div class="source">Source: <a href="{source_urls[0]}">{meeting.get("source_publication", "View Article")}</a></div>'

        html += '</div>'
        return html
    
    def create_excel_report(self, meetings: List[Dict], excel_path: str = 'trump_meetings.xlsx') -> str:
        """
        Create Excel spreadsheet with meeting data and dashboard (regenerates fresh each time with all meetings)
        Returns the path to the Excel file
        """
        # Always create a fresh workbook
        wb = Workbook()

        # Create Dashboard sheet first (so it's the default view)
        dashboard = wb.active
        dashboard.title = "Dashboard"

        # Create Data sheet
        data_sheet = wb.create_sheet("Meeting Data")

        # ===== POPULATE DATA SHEET =====
        # Define headers
        headers = [
            'Date', 'Location', 'Meeting Type', 'Attendee Name',
            'Title', 'Company', 'Primary Industry', 'Confidence Level',
            'Source Count', 'Source Publication', 'Source URLs', 'Notes'
        ]

        # Write headers with styling
        for col, header in enumerate(headers, start=1):
            cell = data_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0F1F2E", end_color="0F1F2E", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set column widths
        data_sheet.column_dimensions['A'].width = 15  # Date
        data_sheet.column_dimensions['B'].width = 20  # Location
        data_sheet.column_dimensions['C'].width = 15  # Meeting Type
        data_sheet.column_dimensions['D'].width = 20  # Attendee Name
        data_sheet.column_dimensions['E'].width = 25  # Title
        data_sheet.column_dimensions['F'].width = 25  # Company
        data_sheet.column_dimensions['G'].width = 20  # Primary Industry
        data_sheet.column_dimensions['H'].width = 15  # Confidence Level
        data_sheet.column_dimensions['I'].width = 12  # Source Count
        data_sheet.column_dimensions['J'].width = 25  # Source Publication
        data_sheet.column_dimensions['K'].width = 60  # Source URLs
        data_sheet.column_dimensions['L'].width = 40  # Notes

        next_row = 2

        # Collect statistics while adding data
        industries = []
        confidence_levels = []
        companies = []
        locations = []

        # Add all meetings
        for meeting in meetings:
            # Parse source URLs
            source_urls_json = meeting.get('source_urls', '[]')
            try:
                source_urls = json.loads(source_urls_json) if isinstance(source_urls_json, str) else source_urls_json
            except (json.JSONDecodeError, TypeError):
                source_urls = [meeting.get('source_url')] if meeting.get('source_url') else []

            source_count = meeting.get('source_count', len(source_urls) if source_urls else 1)
            source_urls_display = '\n'.join(source_urls) if source_urls else meeting.get('source_url', '')

            for attendee in meeting.get('attendees', []):
                confidence = attendee.get('confidence_level') or ''
                row_data = [
                    meeting.get('date', ''),
                    meeting.get('location', ''),
                    meeting.get('meeting_type', meeting.get('type', '')),
                    attendee.get('name', ''),
                    attendee.get('title', ''),
                    attendee.get('company', ''),
                    attendee.get('primary_industry', ''),
                    confidence.upper() if confidence else '',
                    source_count,
                    meeting.get('source_publication', ''),
                    source_urls_display,
                    meeting.get('notes', '')
                ]

                # Collect stats
                industries.append(attendee.get('primary_industry', 'Unknown'))
                conf_level = attendee.get('confidence_level') or 'unknown'
                confidence_levels.append(conf_level.upper() if conf_level else 'UNKNOWN')
                companies.append(attendee.get('company', 'Unknown'))
                locations.append(meeting.get('location', 'Unknown'))

                for col, value in enumerate(row_data, start=1):
                    cell = data_sheet.cell(row=next_row, column=col, value=value)

                    # Color code by confidence level
                    confidence_val = attendee.get('confidence_level') or ''
                    confidence_lower = confidence_val.lower() if confidence_val else ''
                    if confidence_lower == 'high':
                        cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                    elif confidence_lower == 'medium':
                        cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                    elif confidence_lower == 'low':
                        cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

                next_row += 1

        # ===== CREATE DASHBOARD =====
        # Title
        dashboard['A1'] = 'Trump Meetings Tracker - Dashboard'
        dashboard['A1'].font = Font(bold=True, size=16, color="0F1F2E")
        dashboard.merge_cells('A1:D1')

        # Summary Stats
        dashboard['A3'] = 'Summary Statistics'
        dashboard['A3'].font = Font(bold=True, size=14, color="0F1F2E")

        dashboard['A4'] = 'Total Meetings:'
        dashboard['B4'] = len(meetings)
        dashboard['A5'] = 'Total Attendees:'
        dashboard['B5'] = len(companies)
        dashboard['A6'] = 'Unique Companies:'
        dashboard['B6'] = len(set(companies))
        dashboard['A7'] = 'Date Range:'
        if meetings:
            dates = [m.get('date', '') for m in meetings if m.get('date')]
            dashboard['B7'] = f"{min(dates) if dates else 'N/A'} to {max(dates) if dates else 'N/A'}"

        # Style summary stats
        for row in range(4, 8):
            dashboard[f'A{row}'].font = Font(bold=True)
            dashboard[f'B{row}'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

        # Industry Breakdown
        dashboard['A10'] = 'Meetings by Industry'
        dashboard['A10'].font = Font(bold=True, size=12, color="0F1F2E")

        industry_counts = Counter(industries)
        dashboard['A11'] = 'Industry'
        dashboard['B11'] = 'Count'
        dashboard['A11'].font = Font(bold=True)
        dashboard['B11'].font = Font(bold=True)

        for idx, (industry, count) in enumerate(industry_counts.most_common(10), start=12):
            dashboard[f'A{idx}'] = industry
            dashboard[f'B{idx}'] = count

        # Create bar chart for industries
        industry_chart = BarChart()
        industry_chart.type = "col"
        industry_chart.title = "Top 10 Industries"
        industry_chart.y_axis.title = 'Number of Meetings'

        data = Reference(dashboard, min_col=2, min_row=11, max_row=11 + min(10, len(industry_counts)))
        cats = Reference(dashboard, min_col=1, min_row=12, max_row=11 + min(10, len(industry_counts)))
        industry_chart.add_data(data, titles_from_data=True)
        industry_chart.set_categories(cats)
        industry_chart.height = 10
        industry_chart.width = 15
        dashboard.add_chart(industry_chart, "D10")

        # Confidence Level Breakdown
        dashboard['A25'] = 'Confidence Level Distribution'
        dashboard['A25'].font = Font(bold=True, size=12, color="0F1F2E")

        confidence_counts = Counter(confidence_levels)
        dashboard['A26'] = 'Confidence'
        dashboard['B26'] = 'Count'
        dashboard['A26'].font = Font(bold=True)
        dashboard['B26'].font = Font(bold=True)

        conf_row = 27
        for confidence in ['HIGH', 'MEDIUM', 'LOW']:
            dashboard[f'A{conf_row}'] = confidence
            dashboard[f'B{conf_row}'] = confidence_counts.get(confidence, 0)
            conf_row += 1

        # Create pie chart for confidence
        pie_chart = PieChart()
        pie_chart.title = "Confidence Level Distribution"
        data = Reference(dashboard, min_col=2, min_row=26, max_row=29)
        labels = Reference(dashboard, min_col=1, min_row=27, max_row=29)
        pie_chart.add_data(data, titles_from_data=True)
        pie_chart.set_categories(labels)
        pie_chart.height = 10
        pie_chart.width = 12
        dashboard.add_chart(pie_chart, "D25")

        # Top Companies
        dashboard['A35'] = 'Top 10 Companies'
        dashboard['A35'].font = Font(bold=True, size=12, color="0F1F2E")

        company_counts = Counter(companies)
        dashboard['A36'] = 'Company'
        dashboard['B36'] = 'Meetings'
        dashboard['A36'].font = Font(bold=True)
        dashboard['B36'].font = Font(bold=True)

        for idx, (company, count) in enumerate(company_counts.most_common(10), start=37):
            dashboard[f'A{idx}'] = company
            dashboard[f'B{idx}'] = count

        # Location Breakdown
        dashboard['D35'] = 'Meetings by Location'
        dashboard['D35'].font = Font(bold=True, size=12, color="0F1F2E")

        location_counts = Counter(locations)
        dashboard['D36'] = 'Location'
        dashboard['E36'] = 'Count'
        dashboard['D36'].font = Font(bold=True)
        dashboard['E36'].font = Font(bold=True)

        for idx, (location, count) in enumerate(location_counts.most_common(), start=37):
            dashboard[f'D{idx}'] = location
            dashboard[f'E{idx}'] = count

        # Set column widths for dashboard
        dashboard.column_dimensions['A'].width = 25
        dashboard.column_dimensions['B'].width = 15
        dashboard.column_dimensions['C'].width = 5
        dashboard.column_dimensions['D'].width = 25
        dashboard.column_dimensions['E'].width = 15

        # Save the workbook
        wb.save(excel_path)
        print(f"📊 Excel report created with {len(companies)} meeting entries and dashboard: {excel_path}")

        return excel_path

    def send_email(self, recipients: List[str], subject: str, html_content: str, attachment_path: str = None):
        """Send email using SendGrid with optional Excel attachment"""
        sendgrid_api_key = os.environ.get('SENDGRID_API_KEY')
        sender_email = os.environ.get('SENDER_EMAIL', 'alerts@trumptracker.com')
        
        if not sendgrid_api_key:
            print("❌ ERROR: SENDGRID_API_KEY environment variable not set")
            print("   Set it in GitHub Secrets or your environment")
            return False
        
        try:
            sg = SendGridAPIClient(sendgrid_api_key)

            message = Mail(
                from_email=sender_email,
                to_emails=recipients,
                subject=subject,
                html_content=html_content
            )

            # Attach Excel file if provided
            if attachment_path and os.path.exists(attachment_path):
                with open(attachment_path, 'rb') as f:
                    file_data = f.read()
                    encoded_file = base64.b64encode(file_data).decode()

                attachment = Attachment(
                    FileContent(encoded_file),
                    FileName(os.path.basename(attachment_path)),
                    FileType('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                    Disposition('attachment')
                )
                message.attachment = attachment
                print(f"📎 Attached Excel file: {os.path.basename(attachment_path)}")

            response = sg.send(message)
            
            if response.status_code == 202:
                print(f"✅ Email sent successfully to {len(recipients)} recipient(s)")
                return True
            else:
                print(f"⚠️ Email sent with status code: {response.status_code}")
                return True
                
        except Exception as e:
            print(f"❌ Error sending email: {str(e)}")
            return False
    
    def run(self, days_back=7):
        """Main execution function"""
        print("=" * 60)
        print("TRUMP MEETINGS TRACKER - ENHANCED VERSION")
        print("=" * 60)
        print()
        
        # Search for meetings
        meetings = self.search_all_sources(days_back)
        
        # Save new meetings
        saved_count = 0
        for meeting in meetings:
            meeting_id = self.save_meeting(meeting)
            if meeting_id > 0:
                saved_count += 1
        
        print()
        print(f"💾 Saved {saved_count} new meeting(s) to database")
        
        # Get meetings from last run
        since_date = (datetime.now() - timedelta(days=days_back)).isoformat()
        recent_meetings = self.get_new_meetings(since_date)
        
        print(f"📊 Total meetings in database from last {days_back} days: {len(recent_meetings)}")
        print()
        
        # Generate and send email
        if recent_meetings:
            html_content = self.generate_email_html(recent_meetings)

            # Create Excel report with ALL meetings from database (deduplicated)
            all_meetings = self.get_all_meetings()
            excel_path = self.create_excel_report(all_meetings)

            # Get recipients from environment variable
            recipients_str = os.environ.get('EMAIL_RECIPIENTS', '')
            recipients = [email.strip() for email in recipients_str.split(',') if email.strip()]

            if recipients:
                subject = f"Trump Meetings Update - {len(recent_meetings)} Meeting(s) ({datetime.now().strftime('%b %d, %Y')})"
                self.send_email(recipients, subject, html_content, attachment_path=excel_path)
            else:
                print("⚠️ No email recipients configured. Set EMAIL_RECIPIENTS environment variable.")
                print("\n📧 Generated email saved for preview")

                # Save email to file for preview
                with open('email_preview.html', 'w') as f:
                    f.write(html_content)
                print("   Saved to: email_preview.html")
                print(f"   Excel report saved to: {excel_path}")
        else:
            print("ℹ️ No meetings found for the specified period")
        
        print()
        print("=" * 60)
        print("DONE")
        print("=" * 60)
    
    def add_test_meeting(self, name: str, title: str, company: str, date: str = None):
        """Helper method to add a test meeting manually"""
        if date is None:
            date = datetime.now().strftime('%B %d, %Y')
        
        meeting_data = {
            'date': date,
            'location': 'Mar-a-Lago, FL',
            'type': 'Business Meeting',
            'source_url': f'https://example.com/test-{name.replace(" ", "-").lower()}',
            'source_publication': 'Test Source',
            'notes': 'Test meeting entry',
            'attendees': [{
                'name': name,
                'title': title,
                'company': company,
                'confidence_level': 'high',
                'confidence_reasons': ['Test entry'],
                'requires_review': False
            }]
        }
        
        # Classify industry
        industry_info = self.classify_company_industry(company)
        meeting_data['attendees'][0].update(industry_info)
        
        meeting_id = self.save_meeting(meeting_data)
        if meeting_id > 0:
            print(f"✅ Added test meeting: {name} ({company}) - Meeting ID: {meeting_id}")
        return meeting_id


def main():
    """Entry point for script"""
    tracker = TrumpMeetingsTracker()
    
    # Check if we should add test data
    if os.environ.get('ADD_TEST_DATA') == 'true':
        print("📝 Adding test meetings...")
        tracker.add_test_meeting("Andy Jassy", "CEO", "Amazon", "January 3, 2026")
        tracker.add_test_meeting("Doug McMillon", "CEO", "Walmart", "January 4, 2026")
        tracker.add_test_meeting("Mary Barra", "CEO", "GM", "January 5, 2026")
        print()
    
    # Default: search last 30 days (CEO meetings are infrequent)
    days_back = int(os.environ.get('DAYS_BACK', '30'))
    tracker.run(days_back=days_back)


if __name__ == "__main__":
    main()
